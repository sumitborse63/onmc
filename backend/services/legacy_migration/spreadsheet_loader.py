"""
Spreadsheet Loader for CSV and Excel files.
No OCR needed — reads files directly and normalizes into LegacyMaterialRecord.
"""

import os
import io
import logging
from typing import List, Dict, Any, Optional, Tuple

import pandas as pd

from .models import LegacyMaterialRecord
from .field_extractor import extract_records_from_table
from .industrial_dictionary import normalize_column_name

logger = logging.getLogger(__name__)


def load_csv(file_path: str, source_filename: str = "") -> Tuple[List[LegacyMaterialRecord], Dict[str, Any]]:
    """
    Load a CSV file and convert to LegacyMaterialRecord list.
    Returns (records, metadata).
    """
    try:
        # Try different encodings
        for encoding in ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']:
            try:
                df = pd.read_csv(file_path, encoding=encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            df = pd.read_csv(file_path, encoding='utf-8', errors='replace')

        return _dataframe_to_records(df, source_filename or os.path.basename(file_path))

    except Exception as e:
        logger.error(f"CSV load failed: {e}")
        return [], {"error": str(e)}


def load_csv_from_bytes(file_bytes: bytes, source_filename: str = "") -> Tuple[List[LegacyMaterialRecord], Dict[str, Any]]:
    """Load CSV from bytes (uploaded file)."""
    try:
        for encoding in ['utf-8', 'latin-1', 'cp1252']:
            try:
                df = pd.read_csv(io.BytesIO(file_bytes), encoding=encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            df = pd.read_csv(io.BytesIO(file_bytes), encoding='utf-8', errors='replace')

        return _dataframe_to_records(df, source_filename)
    except Exception as e:
        logger.error(f"CSV bytes load failed: {e}")
        return [], {"error": str(e)}


def load_excel(file_path: str, source_filename: str = "", sheet_name: Optional[str] = None) -> Tuple[List[LegacyMaterialRecord], Dict[str, Any]]:
    """
    Load an Excel file (XLS/XLSX) and convert to LegacyMaterialRecord list.
    Supports multiple sheets, merged cells, non-row-1 headers.
    """
    try:
        import openpyxl

        # Get sheet names
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        available_sheets = wb.sheetnames
        wb.close()

        target_sheet = sheet_name or available_sheets[0]

        df = pd.read_excel(file_path, sheet_name=target_sheet, engine='openpyxl')

        records, meta = _dataframe_to_records(df, source_filename or os.path.basename(file_path))
        meta["available_sheets"] = available_sheets
        meta["processed_sheet"] = target_sheet

        return records, meta

    except ImportError:
        # Fallback for .xls files
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name or 0, engine='xlrd')
            return _dataframe_to_records(df, source_filename or os.path.basename(file_path))
        except Exception as e2:
            logger.error(f"Excel load failed: {e2}")
            return [], {"error": str(e2)}
    except Exception as e:
        logger.error(f"Excel load failed: {e}")
        return [], {"error": str(e)}


def load_excel_from_bytes(file_bytes: bytes, source_filename: str = "", sheet_name: Optional[str] = None) -> Tuple[List[LegacyMaterialRecord], Dict[str, Any]]:
    """Load Excel from bytes (uploaded file)."""
    try:
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name or 0, engine='openpyxl')
        return _dataframe_to_records(df, source_filename)
    except Exception as e:
        logger.error(f"Excel bytes load failed: {e}")
        return [], {"error": str(e)}


def get_excel_sheets(file_bytes: bytes) -> List[str]:
    """Get list of sheet names from an Excel file."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
        sheets = wb.sheetnames
        wb.close()
        return sheets
    except Exception:
        return []


def _dataframe_to_records(df: pd.DataFrame, source_filename: str) -> Tuple[List[LegacyMaterialRecord], Dict[str, Any]]:
    """Convert a pandas DataFrame to LegacyMaterialRecord list."""

    # Clean the dataframe
    df = _clean_dataframe(df)

    if df.empty:
        return [], {"error": "Empty dataframe after cleaning", "total_rows": 0}

    # Get header and data
    headers = [str(col).strip() for col in df.columns.tolist()]
    data_rows = []
    for _, row in df.iterrows():
        data_rows.append([str(v).strip() if pd.notna(v) else "" for v in row.tolist()])

    # Use the field extractor to build records
    records = extract_records_from_table(
        header_row=headers,
        data_rows=data_rows,
        source_file=source_filename,
        source_page=1,
        ocr_confidence=0.99,  # Direct read = high confidence
    )

    metadata = {
        "total_rows": len(data_rows),
        "columns_detected": headers,
        "records_extracted": len(records),
        "source_type": "SPREADSHEET",
    }

    return records, metadata


def _clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Clean a dataframe: remove empty rows, detect header rows, etc."""

    # Drop completely empty rows and columns
    df = df.dropna(how='all')
    df = df.dropna(axis=1, how='all')

    if df.empty:
        return df

    # Check if first row looks like data rather than headers
    # (if columns are just 0, 1, 2, ... then first row might be headers)
    if all(isinstance(col, (int, float)) for col in df.columns):
        # First row is likely the header
        df.columns = [str(v).strip() if pd.notna(v) else f"col_{i}" for i, v in enumerate(df.iloc[0])]
        df = df.iloc[1:].reset_index(drop=True)

    # Check if named columns are actually "Unnamed: X" (auto-generated)
    unnamed_count = sum(1 for c in df.columns if str(c).startswith('Unnamed'))
    if unnamed_count > len(df.columns) * 0.5:
        # Try to find the actual header row
        for idx in range(min(5, len(df))):
            row_vals = df.iloc[idx].tolist()
            non_null = [v for v in row_vals if pd.notna(v) and str(v).strip()]
            if len(non_null) >= len(df.columns) * 0.5:
                # This row looks like headers
                df.columns = [str(v).strip() if pd.notna(v) else f"col_{i}" for i, v in enumerate(row_vals)]
                df = df.iloc[idx + 1:].reset_index(drop=True)
                break

    # Remove rows that look like repeated headers or section dividers
    df = df[~df.apply(lambda row: _is_divider_row(row), axis=1)]

    return df.reset_index(drop=True)


def _is_divider_row(row) -> bool:
    """Check if a row is a section divider or repeated header."""
    non_null = [str(v).strip() for v in row if pd.notna(v) and str(v).strip()]
    if not non_null:
        return True

    # If all values are dashes, equals, or very short
    if all(len(v) <= 3 and v in ['-', '--', '---', '=', '==', '===', '*', '**'] for v in non_null):
        return True

    return False
